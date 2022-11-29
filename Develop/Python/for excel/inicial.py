'''
primero se debe instalar la libreria openpyxl puego se puede usar con import
pip install openpyxl


el libro es el primer objeto, abre uno existente y asigna a variable una hoja
WORKBOOK -> WORKSHEET -> CELL -> a su vez, desde aqui se llega a las propiedades de cell:
row, column, value, fill, etc
'''

import openpyxl

wb=openpyxl.load_workbook(r"C:\Users\adan\Desktop\Repositorio\Nueva carpeta\know\Develop\Python\for excel\model_object.xlsx")
ws=wb["Sheet1"]




'''LEER VALORES DE LA HOJA'''

'''
print(ws["a10"].value)
print(ws.cell(1,10).value)

print(ws.max_row) #cantidad de rows que contienen valores
print(ws.max_column)
'''

'''todas las celdas de la columna A
for cell in ws['B']:
    print(cell.value)
'''
    
for i in range(1,ws.max_row+1): # primero establece el rango de celdas
    for cell in ws[i]:          # asigna su
        print(cell.value)

